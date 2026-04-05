"""
Generate SubsidyAI data from real ISS Excel export (36,651 applications).
Creates EmulatedEntity, Applicant, Application, runs scoring, creates decisions/payments.
"""
import random
import logging
from datetime import datetime, date, timedelta
from decimal import Decimal

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from faker import Faker

from apps.emulator.models import EmulatedEntity, RFIDMonitoring
from apps.scoring.models import (
    Applicant, Application, ApplicationPeriod, AuditLog, Budget, Decision,
    HardFilterResult, Notification, Payment, Score, ScoreFactor,
    SubsidyDirection, SubsidyType, UserProfile,
)

fake = Faker('ru_RU')
logger = logging.getLogger(__name__)

# ============================================================
# CONSTANTS
# ============================================================

DIRECTION_MAP = {
    'Субсидирование в скотоводстве': 'cattle_meat',
    'Субсидирование в птицеводстве': 'poultry',
    'Субсидирование в овцеводстве': 'sheep',
    'Субсидирование в верблюдоводстве': 'camel',
    'Субсидирование в коневодстве': 'horse',
    'Субсидирование затрат по искусственному осеменению': 'insemination',
    'Субсидирование в пчеловодстве': 'beekeeping',
    'Субсидирование в свиноводстве': 'pig',
    'Субсидирование в козоводстве': 'goat',
}

ANIMAL_TYPE_MAP = {
    'cattle_meat': 'cattle', 'cattle_dairy': 'cattle', 'insemination': 'cattle',
    'sheep': 'sheep', 'goat': 'sheep',
    'horse': 'horse', 'camel': 'camel',
    'pig': 'pig', 'poultry': 'poultry', 'beekeeping': 'cattle',
}

BREEDS = {
    'cattle': ['Ангусская', 'Герефордская', 'Казахская белоголовая', 'Шароле', 'Лимузинская', 'Голштинская', 'Симментальская'],
    'sheep': ['Эдильбаевская', 'Казахская тонкорунная', 'Каракульская', 'Меринос', 'Гиссарская'],
    'horse': ['Казахская', 'Кустанайская', 'Мугалжарская', 'Донская'],
    'camel': ['Казахский бактриан', 'Дромедар'],
    'pig': ['Крупная белая', 'Ландрас', 'Дюрок'],
    'poultry': ['Кросс Ломанн', 'Кросс Хайсекс', 'Бройлер Кобб-500'],
}

STATUS_RISK_MAP = {
    'Исполнена': lambda: 'clean',
    'Одобрена': lambda: random.choices(['clean', 'minor_issues'], [90, 10])[0],
    'Сформировано поручение': lambda: 'clean',
    'Отклонена': lambda: random.choices(['risky', 'fraudulent'], [60, 40])[0],
    'Отозвано': lambda: 'minor_issues',
    'Получена': lambda: random.choices(['clean', 'minor_issues', 'risky'], [70, 20, 10])[0],
}

FARM_PREFIXES = ['ТОО', 'КХ', 'ТОО']
FARM_NAMES = [
    'Агрофирма', 'Байтерек', 'Дала', 'Жулдыз', 'Нур', 'Степь', 'Асыл', 'Тулпар',
    'Береке', 'Достык', 'Шанырак', 'Жайлау', 'Кокше', 'Арман', 'Мерей', 'Алтын',
    'Саулет', 'Кенже', 'Акбота', 'Самал', 'Отрар', 'Сарыарка', 'Табиғат', 'Бастау',
]


def generate_iin():
    y = random.randint(60, 99)
    m = random.randint(1, 12)
    d = random.randint(1, 28)
    rest = random.randint(100000, 999999)
    return f'{y:02d}{m:02d}{d:02d}{rest}'


def generate_bin():
    y = random.randint(10, 25)
    m = random.randint(1, 12)
    rest = random.randint(10000000, 99999999)
    return f'{y:02d}{m:02d}{rest}'


# ============================================================
# JSON GENERATORS
# ============================================================

def gen_giss(risk, amount):
    prev = int(amount * random.uniform(5, 15))
    if risk == 'clean':
        growth = round(random.uniform(3, 18), 2)
    elif risk == 'minor_issues':
        growth = round(random.uniform(-3, 8), 2)
    elif risk == 'risky':
        growth = round(random.uniform(-15, 0), 2)
    else:
        growth = round(random.uniform(-25, -5), 2)

    before = int(prev / (1 + growth / 100)) if growth != -100 else prev
    total_subs = int(amount * random.uniform(1, 5))
    consecutive = 0 if risk == 'clean' else (random.choice([0, 1, 2]) if risk != 'fraudulent' else random.choice([2, 3]))

    return {
        'registered': risk != 'fraudulent' or random.random() > 0.3,
        'gross_production_previous_year': prev,
        'gross_production_year_before': before,
        'growth_rate': growth,
        'obligations_met': risk in ('clean', 'minor_issues'),
        'obligations_required': total_subs >= 100_000_000,
        'total_subsidies_received': total_subs,
        'consecutive_decline_years': consecutive,
        'repeat_violation': risk == 'fraudulent' and random.random() > 0.5,
        'blocked': risk == 'fraudulent' and random.random() > 0.5,
        'block_reason': None,
    }


def gen_ias_rszh(risk, amount, quantity, region, district, entity_type, subsidy_name):
    count = random.randint(1, 4) if risk in ('clean', 'minor_issues') else random.randint(0, 2)
    history = []
    for y in range(count):
        year = 2024 - y
        met = risk in ('clean', 'minor_issues') or random.random() > 0.6
        history.append({
            'year': year,
            'type': subsidy_name[:60],
            'amount': int(amount * random.uniform(0.5, 1.5)),
            'heads': max(1, int(quantity * random.uniform(0.5, 1.2))),
            'status': 'executed' if met else 'pending',
            'obligations_met': met,
        })

    return {
        'registered': risk != 'fraudulent' or random.random() > 0.2,
        'registration_date': f'{random.randint(2015, 2022)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}',
        'entity_type': entity_type,
        'region': region,
        'district': district,
        'subsidy_history': history,
        'total_subsidies_history': sum(h['amount'] for h in history),
        'pending_returns': 0 if risk == 'clean' else random.randint(0, 5),
        'ibspr_registered': risk != 'fraudulent' or random.random() > 0.3,
    }


def gen_easu(entity_type, name):
    return {
        'has_account_number': True,
        'account_numbers': [f'KZ-AGR-{random.randint(100000, 999999)}'],
        'is_spk': entity_type == 'cooperative',
        'spk_members': [],
        'spk_name': name if entity_type == 'cooperative' else None,
    }


def gen_is_iszh(risk, iin_bin, quantity, direction_code):
    animal_type = ANIMAL_TYPE_MAP.get(direction_code, 'cattle')
    breed_list = BREEDS.get(animal_type, ['Порода'])
    num_animals = min(max(int(quantity), 5), 150)

    animals = []
    for _ in range(num_animals):
        age = random.randint(6, 60)
        prev_sub = risk == 'fraudulent' and random.random() > 0.5
        animals.append({
            'tag_number': f'KZ{random.randint(10000000, 99999999)}',
            'type': animal_type,
            'breed': random.choice(breed_list),
            'sex': random.choice(['female', 'male']),
            'age_months': age,
            'age_valid': 6 <= age <= 72 if risk != 'fraudulent' else random.random() > 0.4,
            'previously_subsidized': prev_sub,
            'pedigree_certificate': random.random() > 0.3,
            'owner_iin_bin': iin_bin,
            'owner_match': True,
        })

    mort_pct = round(
        random.uniform(0, 1.5) if risk == 'clean' else
        random.uniform(0.5, 2.5) if risk == 'minor_issues' else
        random.uniform(1, 4) if risk == 'risky' else
        random.uniform(2, 8), 2
    )

    valid = [a for a in animals if a['age_valid'] and not a['previously_subsidized']]
    return {
        'verified': True,
        'animals': animals,
        'total_verified': len(valid),
        'total_rejected': len(animals) - len(valid),
        'rejection_reasons': [],
        'mortality_data': {
            'records': [{'animal_type': animal_type, 'category': 'adult',
                         'total_count': len(animals),
                         'fallen_count': max(0, int(len(animals) * mort_pct / 100)),
                         'mortality_pct': mort_pct, 'period': '2025'}],
            'total_mortality_pct': mort_pct,
            'reporting_year': 2025,
        },
    }


def gen_is_esf(risk, iin_bin, amount, quantity, subsidy_name):
    inv_amount = int(amount * random.uniform(1.2, 2.5))
    return {
        'invoices': [{
            'esf_number': f'ESF-2025-{random.randint(1000000, 9999999)}',
            'date': f'2025-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}',
            'seller_bin': generate_bin(),
            'seller_name': f'ТОО "{random.choice(FARM_NAMES)}"',
            'buyer_iin_bin': iin_bin,
            'total_amount': inv_amount,
            'items': [{'description': subsidy_name[:50], 'quantity': max(1, int(quantity)),
                        'unit': 'голова', 'unit_price': int(inv_amount / max(1, int(quantity))),
                        'amount': inv_amount}],
            'status': 'confirmed' if risk != 'fraudulent' else random.choice(['confirmed', 'draft']),
            'payment_confirmed': risk in ('clean', 'minor_issues'),
        }],
        'total_amount': inv_amount,
        'invoice_count': 1,
    }


def gen_egkn(risk, quantity, region, district):
    if risk == 'fraudulent' and random.random() > 0.6:
        return {'has_agricultural_land': False, 'plots': [], 'total_agricultural_area': 0, 'pasture_area': 0, 'pasture_zone': 'degraded'}

    land_area = round(max(5, quantity * random.uniform(3, 10)), 1)
    pasture_area = round(land_area * random.uniform(0.4, 0.8), 1)

    return {
        'has_agricultural_land': True,
        'total_agricultural_area': land_area,
        'pasture_area': pasture_area,
        'pasture_zone': 'restored' if risk != 'fraudulent' else 'degraded',
        'plots': [{
            'cadastral_number': f'{random.randint(1, 99):02d}-{random.randint(100, 999)}-{random.randint(100, 999)}-{random.randint(100, 999)}',
            'area_hectares': land_area,
            'purpose': 'сельскохозяйственное назначение',
            'sub_purpose': random.choice(['пастбище', 'пашня', 'сенокос']),
            'region': region, 'district': district,
            'ownership_type': random.choice(['собственность', 'аренда']),
            'registration_date': f'{random.randint(2010, 2023)}-01-01',
        }],
    }


def gen_treasury(risk, excel_status, amount):
    if excel_status in ('Исполнена', 'Сформировано поручение'):
        pay_status = 'completed' if excel_status == 'Исполнена' else 'pending'
        return {'payments': [{
            'payment_id': f'PAY-2025-{random.randint(10000, 99999)}',
            'amount': int(amount),
            'status': pay_status,
            'paid_date': f'2025-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}' if pay_status == 'completed' else None,
            'treasury_reference': f'TR-2025-{random.randint(10000, 99999)}',
        }]}
    return {'payments': []}


# ============================================================
# MAIN COMMAND
# ============================================================

class Command(BaseCommand):
    help = 'Generate data from real ISS Excel export (36,651 applications)'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, required=True, help='Path to Excel file')
        parser.add_argument('--clear', action='store_true', help='Clear all existing data first')
        parser.add_argument('--skip-shap', action='store_true', help='Skip SHAP calculations (5x faster)')
        parser.add_argument('--batch-size', type=int, default=1000, help='Bulk create batch size')
        parser.add_argument('--limit', type=int, default=0, help='Limit number of rows (0=all)')

    def handle(self, *args, **options):
        filepath = options['file']
        batch_size = options['batch_size']
        skip_shap = options['skip_shap']
        limit = options['limit']

        self.stdout.write(f'Reading Excel: {filepath}')

        # ---- Step 0: Parse Excel ----
        rows = self._parse_excel(filepath, limit)
        self.stdout.write(self.style.SUCCESS(f'Parsed {len(rows)} rows'))

        # ---- Step 0.5: Clear old data ----
        if options['clear']:
            self._clear_data()

        # ---- Step 1: Create directions & subsidy types ----
        directions, subsidy_types = self._create_references(rows)

        # ---- Step 2: Create EmulatedEntity + Applicant + Application ----
        self._create_entities_and_applications(rows, directions, subsidy_types, batch_size)

        # ---- Step 3: Run scoring ----
        self._run_scoring(skip_shap)

        # ---- Step 4: Post-scoring — set real statuses, create decisions/payments ----
        self._post_scoring(rows)

        # ---- Step 5: Create budgets ----
        self._create_budgets(rows, directions)

        # ---- Step 6: Stats ----
        self._print_stats()

    # ----------------------------------------------------------
    # PARSE EXCEL
    # ----------------------------------------------------------
    def _parse_excel(self, filepath, limit):
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] is None or str(row[0]).strip() == '':
                continue

            rate = float(row[10]) if row[10] else 0
            amount = float(row[11]) if row[11] else 0
            if rate <= 0 or amount <= 0:
                continue

            # Parse date
            date_str = str(row[1]).strip() if row[1] else ''
            try:
                submitted_dt = datetime.strptime(date_str, '%d.%m.%Y %H:%M:%S')
            except (ValueError, TypeError):
                try:
                    submitted_dt = datetime.strptime(date_str, '%d.%m.%Y %H:%M')
                except (ValueError, TypeError):
                    submitted_dt = datetime(2025, 1, 1, 12, 0, 0)

            rows.append({
                'row_num': int(row[0]) if row[0] else 0,
                'submitted_dt': submitted_dt,
                'region': str(row[4]).strip() if row[4] else '',
                'akimat': str(row[5]).strip() if row[5] else '',
                'app_number': str(row[6]).strip() if row[6] else f'AUTO-{len(rows)}',
                'direction_name': str(row[7]).strip() if row[7] else '',
                'subsidy_name': str(row[8]).strip() if row[8] else '',
                'excel_status': str(row[9]).strip() if row[9] else '',
                'rate': rate,
                'amount': amount,
                'district': str(row[12]).strip() if row[12] else '',
            })

            if limit and len(rows) >= limit:
                break

        wb.close()
        return rows

    # ----------------------------------------------------------
    # CLEAR DATA
    # ----------------------------------------------------------
    def _clear_data(self):
        self.stdout.write('Clearing old data...')
        counts = {}
        for model in [AuditLog, Notification, Payment, Decision, ScoreFactor, Score,
                       HardFilterResult, Application, Applicant, RFIDMonitoring,
                       EmulatedEntity, Budget]:
            name = model.__name__
            deleted, _ = model.objects.all().delete()
            counts[name] = deleted

        for name, cnt in counts.items():
            if cnt > 0:
                self.stdout.write(f'  Deleted {cnt} {name}')
        self.stdout.write(self.style.SUCCESS('Data cleared'))

    # ----------------------------------------------------------
    # CREATE REFERENCES
    # ----------------------------------------------------------
    def _create_references(self, rows):
        self.stdout.write('Creating directions and subsidy types...')

        # Directions
        directions = {}
        for dname, dcode in DIRECTION_MAP.items():
            obj, _ = SubsidyDirection.objects.get_or_create(
                code=dcode, defaults={'name': dname, 'is_active': True}
            )
            directions[dname] = obj

        # Subsidy types — unique (direction_name, subsidy_name, rate)
        seen_types = set()
        subsidy_types = {}
        form_counter = {}

        for r in rows:
            key = (r['direction_name'], r['subsidy_name'], r['rate'])
            if key in seen_types:
                continue
            seen_types.add(key)

            direction = directions.get(r['direction_name'])
            if not direction:
                continue

            # Determine form_number from prefix
            prefix = r['app_number'][:4] if r['app_number'] else '0000'
            form_num = form_counter.get(r['direction_name'], 0) + 1
            form_counter[r['direction_name']] = form_num

            obj, _ = SubsidyType.objects.get_or_create(
                direction=direction,
                name=r['subsidy_name'][:500],
                rate=Decimal(str(r['rate'])),
                defaults={
                    'form_number': form_num,
                    'unit': 'голова',
                    'origin': 'domestic',
                    'is_active': True,
                },
            )
            subsidy_types[key] = obj

        self.stdout.write(f'  Directions: {len(directions)}, SubsidyTypes: {len(subsidy_types)}')
        return directions, subsidy_types

    # ----------------------------------------------------------
    # CREATE ENTITIES AND APPLICATIONS
    # ----------------------------------------------------------
    def _create_entities_and_applications(self, rows, directions, subsidy_types, batch_size):
        self.stdout.write(f'Creating {len(rows)} entities and applications...')

        existing_iins = set()
        entities_batch = []
        applicants_batch = []
        applications_batch = []

        for i, r in enumerate(rows):
            direction = directions.get(r['direction_name'])
            st_key = (r['direction_name'], r['subsidy_name'], r['rate'])
            subsidy_type = subsidy_types.get(st_key)
            if not direction or not subsidy_type:
                continue

            quantity = max(1, int(r['amount'] / r['rate']))
            risk = STATUS_RISK_MAP.get(r['excel_status'], lambda: 'clean')()
            direction_code = DIRECTION_MAP.get(r['direction_name'], 'cattle_meat')

            # Entity type by amount
            if r['amount'] > 5_000_000:
                entity_type = random.choice(['legal', 'legal', 'cooperative'])
            elif r['amount'] > 500_000:
                entity_type = random.choice(['legal', 'individual'])
            else:
                entity_type = 'individual'

            # Unique IIN/BIN
            while True:
                iin_bin = generate_bin() if entity_type != 'individual' else generate_iin()
                if iin_bin not in existing_iins:
                    existing_iins.add(iin_bin)
                    break

            # Name
            if entity_type == 'cooperative':
                name = f'СПК "{random.choice(FARM_NAMES)}"'
            elif entity_type == 'legal':
                name = f'{random.choice(FARM_PREFIXES)} "{random.choice(FARM_NAMES)} {r["district"][:15]}"'
            else:
                name = f'ИП {fake.last_name()} {fake.first_name()[0]}.'

            # EmulatedEntity
            entity = EmulatedEntity(
                iin_bin=iin_bin,
                name=name,
                entity_type=entity_type,
                region=r['region'],
                district=r['district'],
                registration_date=date(random.randint(2015, 2023), random.randint(1, 12), random.randint(1, 28)),
                risk_profile=risk,
                giss_data=gen_giss(risk, r['amount']),
                ias_rszh_data=gen_ias_rszh(risk, r['amount'], quantity, r['region'], r['district'], entity_type, r['subsidy_name']),
                easu_data=gen_easu(entity_type, name),
                is_iszh_data=gen_is_iszh(risk, iin_bin, quantity, direction_code),
                is_esf_data=gen_is_esf(risk, iin_bin, r['amount'], quantity, r['subsidy_name']),
                egkn_data=gen_egkn(risk, quantity, r['region'], r['district']),
                treasury_data=gen_treasury(risk, r['excel_status'], r['amount']),
            )
            entities_batch.append(entity)

            # Applicant
            applicant = Applicant(
                iin_bin=iin_bin,
                name=name,
                entity_type=entity_type,
                region=r['region'],
                district=r['district'],
                address=f'{r["region"]}, {r["district"]}',
                registration_date=entity.registration_date,
                bank_account=f'KZ{random.randint(10, 99)}{"".join([str(random.randint(0,9)) for _ in range(18)])}',
                bank_name=random.choice(['Халык банк', 'Kaspi банк', 'Forte банк', 'БЦК', 'АТФ банк']),
            )
            applicants_batch.append(applicant)

            # Application — status starts as submitted for scoring, real status set later
            app_status = 'draft' if r['excel_status'] == 'Отозвано' else 'submitted'
            application = Application(
                number=r['app_number'],
                applicant=None,  # will be set after bulk_create
                subsidy_type=subsidy_type,
                status=app_status,
                quantity=quantity,
                unit_price=Decimal(str(int(r['amount'] / quantity * random.uniform(1.5, 3)))),
                rate=Decimal(str(r['rate'])),
                total_amount=Decimal(str(int(r['amount']))),
                submitted_at=timezone.make_aware(r['submitted_dt']),
                region=r['region'],
                district=r['district'],
                akimat=r['akimat'],
                notes=r['excel_status'],  # store real status for post-scoring
            )
            applications_batch.append(application)

            # Bulk create when batch full
            if len(entities_batch) >= batch_size:
                self._flush_batch(entities_batch, applicants_batch, applications_batch, i + 1, len(rows))
                entities_batch = []
                applicants_batch = []
                applications_batch = []

        # Flush remaining
        if entities_batch:
            self._flush_batch(entities_batch, applicants_batch, applications_batch, len(rows), len(rows))

    def _flush_batch(self, entities, applicants, applications, current, total):
        with transaction.atomic():
            # 1. Create entities
            EmulatedEntity.objects.bulk_create(entities, ignore_conflicts=True)

            # 2. Create applicants
            Applicant.objects.bulk_create(applicants, ignore_conflicts=True)

            # 3. Re-fetch applicants by iin_bin to set FK on applications
            iin_map = {a.iin_bin: a for a in Applicant.objects.filter(
                iin_bin__in=[a.iin_bin for a in applicants]
            )}

            for app, ent in zip(applications, entities):
                applicant = iin_map.get(ent.iin_bin)
                if applicant:
                    app.applicant = applicant
                    app.applicant_id = applicant.pk

            # Filter out apps without applicant
            valid_apps = [a for a in applications if a.applicant_id]
            Application.objects.bulk_create(valid_apps, ignore_conflicts=True)

        self.stdout.write(f'  Created batch: {current}/{total}')

    # ----------------------------------------------------------
    # RUN SCORING
    # ----------------------------------------------------------
    def _run_scoring(self, skip_shap):
        from apps.scoring.scoring_engine import ScoringEngine

        apps = Application.objects.filter(status='submitted').select_related(
            'applicant', 'subsidy_type', 'subsidy_type__direction'
        ).order_by('id')

        total = apps.count()
        self.stdout.write(f'Running scoring on {total} applications (skip_shap={skip_shap})...')

        if skip_shap:
            # Monkey-patch to skip SHAP
            from apps.scoring import ml_model
            original_shap = ml_model.explain_with_shap
            ml_model.explain_with_shap = lambda *args, **kwargs: None

        engine = ScoringEngine()
        scored = 0
        errors = 0

        for app in apps.iterator(chunk_size=500):
            try:
                engine.run_scoring(app)
                scored += 1
            except Exception as e:
                errors += 1
                if errors <= 10:
                    self.stderr.write(f'  Error scoring {app.number}: {e}')

            if scored % 500 == 0:
                self.stdout.write(f'  Scored: {scored}/{total} (errors: {errors})')

        if skip_shap:
            ml_model.explain_with_shap = original_shap

        self.stdout.write(self.style.SUCCESS(f'Scoring complete: {scored} scored, {errors} errors'))

    # ----------------------------------------------------------
    # POST-SCORING
    # ----------------------------------------------------------
    def _post_scoring(self, rows):
        self.stdout.write('Setting real statuses, creating decisions and payments...')

        from apps.scoring.scoring_engine import calculate_merit_score

        # Build number → excel_status map
        status_map = {r['app_number']: r['excel_status'] for r in rows}

        # Status mapping
        EXCEL_TO_DB_STATUS = {
            'Исполнена': 'paid',
            'Одобрена': 'approved',
            'Сформировано поручение': 'approved',
            'Отклонена': 'rejected',
            'Отозвано': 'draft',
            'Получена': 'submitted',
        }

        PAYMENT_STATUS_MAP = {
            'Исполнена': 'completed',
            'Одобрена': 'initiated',
            'Сформировано поручение': 'sent_to_treasury',
        }

        decisions_batch = []
        payments_batch = []
        apps_to_update = []

        apps = Application.objects.all().select_related('subsidy_type', 'subsidy_type__direction')
        total = apps.count()
        processed = 0

        for app in apps.iterator(chunk_size=1000):
            excel_status = status_map.get(app.number, app.notes or '')
            new_status = EXCEL_TO_DB_STATUS.get(excel_status, app.status)

            if app.status != new_status:
                app.status = new_status
                apps_to_update.append(app)

            # Decision
            if excel_status in ('Исполнена', 'Одобрена', 'Сформировано поручение'):
                decisions_batch.append(Decision(
                    application=app,
                    decision='approved',
                    decided_by_id=1,  # admin
                    reason=f'Одобрено комиссией. AI-скор: {getattr(app, "_cached_score", "N/A")}',
                    approved_amount=app.total_amount,
                ))
            elif excel_status == 'Отклонена':
                decisions_batch.append(Decision(
                    application=app,
                    decision='rejected',
                    decided_by_id=1,
                    reason='Отклонено комиссией по результатам проверки.',
                ))

            # Payment
            if excel_status in PAYMENT_STATUS_MAP:
                merit, breakdown = calculate_merit_score(app)
                payments_batch.append(Payment(
                    application=app,
                    amount=app.total_amount,
                    status=PAYMENT_STATUS_MAP[excel_status],
                    merit_score=Decimal(str(merit)),
                    merit_breakdown=breakdown,
                    initiated_by_id=1,
                ))

            processed += 1
            if processed % 2000 == 0:
                # Flush
                self._flush_post_scoring(apps_to_update, decisions_batch, payments_batch)
                apps_to_update = []
                decisions_batch = []
                payments_batch = []
                self.stdout.write(f'  Post-scoring: {processed}/{total}')

        # Final flush
        self._flush_post_scoring(apps_to_update, decisions_batch, payments_batch)
        self.stdout.write(self.style.SUCCESS(f'Post-scoring complete: {processed} processed'))

    def _flush_post_scoring(self, apps, decisions, payments):
        with transaction.atomic():
            if apps:
                Application.objects.bulk_update(apps, ['status'], batch_size=1000)
            if decisions:
                Decision.objects.bulk_create(decisions, ignore_conflicts=True)
            if payments:
                Payment.objects.bulk_create(payments, ignore_conflicts=True)

    # ----------------------------------------------------------
    # CREATE BUDGETS
    # ----------------------------------------------------------
    def _create_budgets(self, rows, directions):
        self.stdout.write('Creating budgets...')
        from collections import defaultdict

        budget_data = defaultdict(lambda: {'planned': 0, 'spent': 0})
        for r in rows:
            direction = directions.get(r['direction_name'])
            if not direction:
                continue
            key = (r['region'], direction.pk)
            budget_data[key]['planned'] += r['amount']
            if r['excel_status'] == 'Исполнена':
                budget_data[key]['spent'] += r['amount']

        for (region, dir_pk), data in budget_data.items():
            Budget.objects.update_or_create(
                year=2025, region=region, direction_id=dir_pk,
                defaults={
                    'planned_amount': Decimal(str(int(data['planned'] * 1.2))),
                    'spent_amount': Decimal(str(int(data['spent']))),
                },
            )
        self.stdout.write(f'  Created {len(budget_data)} budgets')

    # ----------------------------------------------------------
    # STATS
    # ----------------------------------------------------------
    def _print_stats(self):
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write('GENERATION RESULTS')
        self.stdout.write('=' * 60)

        stats = {
            'EmulatedEntity': EmulatedEntity.objects.count(),
            'Applicant': Applicant.objects.count(),
            'Application': Application.objects.count(),
            'HardFilterResult': HardFilterResult.objects.count(),
            'Score': Score.objects.count(),
            'ScoreFactor': ScoreFactor.objects.count(),
            'Decision': Decision.objects.count(),
            'Payment': Payment.objects.count(),
            'Budget': Budget.objects.count(),
            'Notification': Notification.objects.count(),
        }

        for name, cnt in stats.items():
            self.stdout.write(f'  {name:<20} {cnt:>10,}')

        # Score distribution
        from django.db.models import Avg, Count
        score_stats = Score.objects.aggregate(
            avg_score=Avg('total_score'),
            total=Count('id'),
        )
        if score_stats['total']:
            self.stdout.write(f'\n  Average AI score: {score_stats["avg_score"]:.1f}')

        rec_counts = Score.objects.values('recommendation').annotate(cnt=Count('id'))
        for r in rec_counts:
            self.stdout.write(f'  {r["recommendation"]}: {r["cnt"]:,}')

        # Application status distribution
        self.stdout.write('\n  Application statuses:')
        status_counts = Application.objects.values('status').annotate(cnt=Count('id'))
        for s in status_counts:
            self.stdout.write(f'    {s["status"]}: {s["cnt"]:,}')

        self.stdout.write(self.style.SUCCESS('\nDone!'))
